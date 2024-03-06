import pandas as pd
import polars as pl
import numpy as np
import os
from openpyxl.styles import PatternFill
from openpyxl.worksheet.filters import FilterColumn, Filters
from openpyxl.utils import get_column_letter
from get_nearest_feature import get_nearest_features

def gene2peak(gene_file, peak_file, peak_type, species, feature_type, num_features, ref_dir, output_name,
              option = 'native_peak_boundaries', boundary = None, num_peaks_cutoff = None):
    if peak_type == 'MACS2' and 'xls' in peak_file:
        peaks = read_input_MACS2_xls(peak_file)
        peaks = process_input_MACS2_xls(peaks)
    elif peak_type == 'MACS2' and 'bed' in peak_file:
        peaks = read_input_MACS2_bed(peak_file)
        peaks = process_input_MACS2_bed(peaks)
    elif peak_type == 'SEACR':
        peaks = read_input_SEACR(peak_file)
        peaks = process_input_SEACR(peaks)
    else:
        raise TypeError('Invalid peak type')
    
    if 'bed' in peak_file:
        peaks['start'] = peaks['start'] + 1
        peaks['end'] = peaks['end'] + 1
    
    genes = pl.read_csv(gene_file, has_header = False).to_numpy()[:, 0].tolist()
    gene_df = pl.DataFrame()
    for csv in os.listdir('reference/' + species + '/gene/'):
        cur = pl.read_csv('reference/' + species + '/gene/' + csv)
        for gene in genes:
            if gene in cur.select(['gene_name']).to_numpy():
                gene_df = pl.concat([gene_df, cur.filter(pl.col("gene_name") == gene)])
                genes.remove(gene)
    
    gene_df = gene_df.rename({'gene_name': 'name'})
    decomposed_peaks = decompose_peaks(peaks)
    decomposed_genes = decompose_genes(gene_df)
    gen_output(decomposed_peaks, decomposed_genes, species, feature_type, num_features, ref_dir, output_name)
    
def read_input_MACS2_xls(file_path):
    peaks = pl.read_csv(file_path, separator = '\t', skip_rows=22)
    peaks = peaks.rename({'-log10(pvalue)': 'neg_log10_pvalue', '-log10(qvalue)': 'neg_log10_qvalue'})
    return peaks

def read_input_MACS2_bed(file_path):
    col_names = ['chr', 'start', 'end', 'name', 'score', 'ph']
    peaks = pl.read_csv(file_path, header = False, separator = '\t', new_columns=col_names)

    return peaks

def read_input_SEACR(file_path):
    col_names = ['chr', 'start', 'end', 'name', 'max_signal', 'max_signal_region']
    peaks = pl.read_csv(file_path, separator = '\t', new_columns=col_names)

    return peaks

def process_input_MACS2_xls(data, qval = None, option = 'native_peak_boundaries', 
                  boundary = None, num_peaks_cutoff = None):
    peaks = data
    if qval is not None:
        peaks = data.select(threshold = pl.when(data.col('neg_log10_qvalue') > -np.log10(qval)))
        peaks = peaks.sort('neg_log10_qvalue', descending=True)

    if num_peaks_cutoff is not None:
        peaks = peaks.head(num_peaks_cutoff)

    peaks = peak_start_end(peaks, option)
    return peaks

def process_input_MACS2_bed(data, score = 0.05, option = 'native_peak_boundaries', 
                  boundary = None, num_peaks_cutoff = None):
    peaks = data

    if num_peaks_cutoff is not None:
        peaks = peaks.head(num_peaks_cutoff)

    peaks = peak_start_end(peaks, option)

    return peaks

def process_input_SEACR(data, signal = None, option = 'native_peak_boundaries',
                        boundary = None, num_peaks_cutoff = None):
    peaks = data
    if signal is not None:
        peaks = data.select(threshold = pl.where(data.col('signal') > signal))
    # peaks.sort_values(by = 'neg_log10_qvalue', ascending=False, inplace=True)

    if num_peaks_cutoff is not None:
        peaks = peaks.head(num_peaks_cutoff)

    peaks = peak_start_end(peaks, option)

    return peaks

def peak_start_end(data, option):
    peaks = data

    if option == 'peak_summit':
        peaks.drop_in_place('start')
        peaks.with_columns((peaks.col('abs_summit')).alias('start'))
        peaks.drop_in_place('end')
        peaks.with_columns((peaks.col('abs_summit')).alias('end'))
    elif option == 'artifical_peak_boundaries' and boundary is not None:
        peaks.drop_in_place('start')
        peaks.with_columns((peaks.col('abs_summit') - boundary).alias('start'))
        peaks.drop_in_place('end')
        peaks.with_columns((peaks.col('abs_summit') + boundary).alias('end'))
    elif option == 'native_peak_boundaries':
        pass
    else:
        raise ValueError('Invalid peak start/end option')
    
    return peaks

def decompose_genes(genes):
    return {str(name[0]): group for name, group in genes.group_by(['chr'])}

def decompose_peaks(peaks):
    return {'chr' + str(name[0]): group for name, group in peaks.group_by(['chr'])}

def gen_output(decomposed_peaks, decomposed_genes, species, feature_type, num_features, ref_dir, output_name):
    output = pl.DataFrame()
    for key in decomposed_genes.keys():
        if key in decomposed_peaks.keys():
            starts = decomposed_peaks[key].select(['name', 'start'])
            ends = decomposed_peaks[key].select(['name', 'end'])
        else:
            starts = pl.DataFrame({}, schema=['name', 'start'])
            ends = pl.DataFrame({}, schema=['name', 'end'])
        output = pl.concat([output, get_nearest_features(decomposed_genes[key], 'name', starts, ends, None, None, num_features)])
            
    if not os.path.exists("results/"):
        os.mkdir("results/")
    
    output = output.to_pandas()
    output = output.sort_values(by=['chr', 'name'])

    with pd.ExcelWriter('results/' + output_name + '.xlsx', engine='openpyxl') as writer:
    
        output.to_excel(writer, sheet_name='Sheet1', index=False)

        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        for row_num in range(2, len(output) + 2):
            if row_num % 2 == 0:
                for col_num in range(1, output.shape[1] + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
            else:
                for col_num in range(1, output.shape[1] + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width


        unique_chr_values = output['chr'].unique()

        filters = worksheet.auto_filter
        filters.ref = "A1:A" + str(len(output) + 1)
        col = FilterColumn(colId=0)
        col.filters = Filters(filter=unique_chr_values.tolist())
        filters.filterColumn.append(col)

        # Save the Excel file
        workbook.save('results/' + output_name + '.xlsx')
 
# data information
data_dir = "test/"
ref_dir = "reference/"

# function parameters
num_peaks_cutoff = None 
num_nearest_features = 3

# Set peak boundary options
# a. "native_peak_boundaries" - use start + end of peak, as defined by peak caller
# b. "peak_summit" - use peak summit
# c. "artificial_peak_boundaries" - use artificial boundary, such as +/-100 bp from peak summit
option = "native_peak_boundaries"
boundary = None

gene2peak(data_dir + 'gene_to_find.csv', data_dir + 'MACS2_peaks.xls', 'MACS2', "mm10", "name", 3, ref_dir, 'gene_to_find')
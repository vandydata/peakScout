import pandas as pd
import numpy as np
import os
from openpyxl.styles import PatternFill
from openpyxl.worksheet.filters import FilterColumn, Filters
from openpyxl.utils import get_column_letter

def peakScout(file_path, peak_type, species, feature_type, num_features, ref_dir, output_name,
              option = 'native_peak_boundaries', boundary = None, num_peaks_cutoff = None):
    if peak_type == 'MACS2' and 'xls' in file_path:
        peaks = read_input_MACS2_xls(file_path)
        peaks = process_input_MACS2_xls(peaks)
    elif peak_type == 'MACS2' and 'bed' in file_path:
        peaks = read_input_MACS2_bed(file_path)
        peaks = process_input_MACS2_bed(peaks)
    elif peak_type == 'SEACR':
        peaks = read_input_SEACR(file_path)
        peaks = process_input_SEACR(peaks)
    else:
        raise TypeError('Invalid peak type')
    
    print(peaks['start'])

    if 'bed' in file_path:
        peaks['start'] = peaks['start'] + 1
        peaks['end'] = peaks['end'] + 1
    
    print(peaks['start'])
    
    decomposed_peaks = decompose_peaks(peaks)
    gen_output(decomposed_peaks, species, feature_type, num_features, ref_dir, output_name)

def read_input_MACS2_xls(file_path):
    peaks = pd.read_csv(file_path, sep = '\t', skiprows=22)
    peaks.rename(columns={'-log10(pvalue)': 'neg_log10_pvalue', '-log10(qvalue)': 'neg_log10_qvalue'}, inplace=True)

    return peaks

def read_input_MACS2_bed(file_path):
    col_names = ['chr', 'start', 'end', 'name', 'score', 'ph']
    peaks = pd.read_csv(file_path, sep = '\t', names=col_names)

    return peaks

def read_input_SEACR(file_path):
    col_names = ['chr', 'start', 'end', 'name', 'max_signal', 'max_signal_region']
    peaks = pd.read_csv(file_path, sep = '\t', names=col_names)

    return peaks

def process_input_MACS2_xls(data, qval = 0.05, option = 'native_peak_boundaries', 
                  boundary = None, num_peaks_cutoff = None):
    
    peaks = data.loc[data['neg_log10_qvalue'] > -np.log10(qval)]
    peaks.sort_values(by = 'neg_log10_qvalue', ascending=False, inplace=True)

    if num_peaks_cutoff is not None:
        peaks = peaks.head(num_peaks_cutoff)

    if option == 'peak_summit':
        peaks['start'] = peaks['abs_summit']
        peaks['end'] = peaks['abs_summit']
    elif option == 'artifical_peak_boundaries' and boundary is not None:
        peaks['start'] = peaks['abs_summit'] - boundary
        peaks['end'] = peaks['abs_summit'] + boundary
    elif option == 'native_peak_boundaries':
        pass
    else:
        raise ValueError('Invalid peak start/end option')
    
    return peaks

def process_input_MACS2_bed(data, score = 0.05, option = 'native_peak_boundaries', 
                  boundary = None, num_peaks_cutoff = None):
    
    peaks = data

    if num_peaks_cutoff is not None:
        peaks = peaks.head(num_peaks_cutoff)

    if option == 'peak_summit':
        peaks['start'] = peaks['abs_summit']
        peaks['end'] = peaks['abs_summit']
    elif option == 'artifical_peak_boundaries' and boundary is not None:
        peaks['start'] = peaks['abs_summit'] - boundary
        peaks['end'] = peaks['abs_summit'] + boundary
    elif option == 'native_peak_boundaries':
        pass
    else:
        raise ValueError('Invalid peak start/end option')
    
    return peaks

def process_input_SEACR(data, signal = None, option = 'native_peak_boundaries',
                        boundary = None, num_peaks_cutoff = None):
    peaks = data
    if signal is not None:
        peaks = data.loc[data['signal'] > signal]
    # peaks.sort_values(by = 'neg_log10_qvalue', ascending=False, inplace=True)

    if num_peaks_cutoff is not None:
        peaks = peaks.head(num_peaks_cutoff)

    if option == 'peak_summit':
        peaks['start'] = peaks['abs_summit']
        peaks['end'] = peaks['abs_summit']
    elif option == 'artifical_peak_boundaries' and boundary is not None:
        peaks['start'] = peaks['abs_summit'] - boundary
        peaks['end'] = peaks['abs_summit'] + boundary
    elif option == 'native_peak_boundaries':
        pass
    else:
        raise ValueError('Invalid peak start/end option')
    
    return peaks

def decompose_peaks(peaks):
    return {'chr' + str(name[0]): group for name, group in peaks.groupby(['chr'], group_keys=False)}

def gen_output(decomposed_peaks, species, feature_type, num_features, ref_dir, output_name):
    output = pd.DataFrame()
    for key in decomposed_peaks.keys():
        starts = pd.read_csv(ref_dir + species + "/" + feature_type + "/" + key + '_start.csv')
        ends = pd.read_csv(ref_dir + species + "/" + feature_type + "/" + key + '_end.csv')
        output = pd.concat([output, get_nearest_features(decomposed_peaks[key], starts, ends, num_features)], 
                        ignore_index = False, sort = False)

    if not os.path.exists("results/"):
        os.mkdir("results/")
        
    with pd.ExcelWriter('results/' + output_name + '.xlsx', engine='openpyxl') as writer:
    
        output.to_excel(writer, sheet_name='Sheet1', index=False)

        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        for row_num in range(2, len(output) + 2):
            if row_num % 2 == 0:
                for col_num in range(1, output.shape[1] + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
            else:
                for col_num in range(1, output.shape[1] + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width


        unique_chr_values = output['chr'].unique()

        filters = worksheet.auto_filter
        filters.ref = "A1:A" + str(len(output) + 1)
        col = FilterColumn(colId=0)
        col.filters = Filters(filter=unique_chr_values.tolist())
        filters.filterColumn.append(col)

        # Save the Excel file
        workbook.save('results/' + output_name + '.xlsx')

def get_nearest_features(roi, starts, ends, k):
    gene_starts = starts['start'].values
    gene_ends = ends['end'].values
    assert(len(gene_starts) == len(gene_ends))

    return_roi = None
    if 'name' in roi.columns:
        return_roi = roi[['chr', 'name', 'start', 'end']].copy()
    else:
        return_roi = roi[['chr', 'start', 'end']].copy()

    for i in range(1, k+1):
        return_roi['closest_gene_' + str(i)] = None
        return_roi['closest_gene_' + str(i) + '_dist'] = None

    index = 0
    for _, peak in return_roi.iterrows():
        peak_start = peak['start']
        peak_end = peak['end']
        downstream_index = gene_starts.searchsorted(peak_start, side='left')
        upstream_index = gene_ends.searchsorted(peak_end, side='right') - 1
        i = k

        while i > 0 and upstream_index > -1 and downstream_index < len(starts):
            downstream_dist = gene_starts[downstream_index] - peak_end
            upstream_dist = peak_start - gene_ends[upstream_index]
            closest = None

            downstream_dist = 0 if downstream_dist < 0 else downstream_dist
            upstream_dist = 0 if upstream_dist < 0 else upstream_dist

            if downstream_dist < upstream_dist:
                closest = downstream_index
                return_roi.iloc[index, len(return_roi.columns) - 2*i + 1] = downstream_dist
                return_roi.iloc[index, len(return_roi.columns) - 2*i] = starts.iloc[closest, :]['gene_name']
                downstream_index += 1
            else:
                closest = upstream_index
                return_roi.iloc[index, len(return_roi.columns) - 2*i + 1] = upstream_dist
                return_roi.iloc[index, len(return_roi.columns) - 2*i] = ends.iloc[closest, :]['gene_name']
                upstream_index -= 1
        
            i -= 1
        
        if i > 0 and upstream_index < 0:
            while i > 0 and downstream_index < len(starts):
                return_roi.iloc[index, len(return_roi.columns) - i] = starts.iloc[downstream_index, :]['gene_name']
                downstream_index += 1
                i -= 1
        elif i > 0 and downstream_index >= len(ends):
            while i > 0 and upstream_index > -1:
                return_roi.iloc[index, len(return_roi.columns) - i] = ends.iloc[upstream_index, :]['gene_name']
                upstream_index -= 1
                i -= 1

        index += 1

    return return_roi
 
# data information
data_dir = "test/"
ref_dir = "reference/"

# function parameters
num_peaks_cutoff = None 
num_nearest_features = 3

# Set peak boundary options
# a. "native_peak_boundaries" - use start + end of peak, as defined by peak caller
# b. "peak_summit" - use peak summit
# c. "artificial_peak_boundaries" - use artificial boundary, such as +/-100 bp from peak summit
option = "native_peak_boundaries"
boundary = None

for file in os.listdir(data_dir):
    # if "macs2" in file or "MACS2" in file:
    # if "MACS2" in file:
    #     peakScout(data_dir + file, 'MACS2', "mm10", "gene", 3, ref_dir, file[:-4])
    if "seacr" in file or "SEACR" in file:
        peakScout(data_dir + file, 'SEACR', "mm10", "gene", 3, ref_dir, file[:-4])
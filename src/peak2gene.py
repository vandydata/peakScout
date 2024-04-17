import pandas as pd
import polars as pl
import os
from openpyxl.styles import PatternFill
from openpyxl.worksheet.filters import FilterColumn, Filters
from openpyxl.utils import get_column_letter
from get_nearest_feature import get_nearest_features
from process_input import process_input

def peak2gene(peak_file: str,
              peak_type: str, 
              species: str, 
              num_features: int, 
              ref_dir: str, 
              output_name: str,
              out_dir: str, 
              output_type: str,
              option: str = 'native_peak_boundaries', 
              boundary: int = None, 
              up_bound: int = None, 
              down_bound: int = None) -> None:
    '''
    Find the nearest genes for a given list of peaks.

    Parameters:
    peak_file (str): Path to the peak file.
    peak_type (str): Type of peak caller used to generate peak file (e.g. MACS2, SEACR).
    species (str): Species of the reference genome.
    num_features (int): Number of nearest features to find.
    ref_dir (str): Directory containing decomposed reference data.
    output_name (str): Name for output file.
    out_dir (str): Directory to output file.
    output_type (str): Output type (csv file or xlsx file)
    option (str): Option for defining start and end positions of peaks.
    boundary (int): Boundary for artificial peak boundary option. None if other options. 
    up_bound (int): Maximum allowed distance between peak and upstream feature.
    down_bound (int): Maximum allowed distance between peak and downstream feature.

    Returns:
    None

    Outputs:
    Excel sheet containing peak data, the nearest k genes for each peak, and the distance
    between those genes and the peak.
    '''
    
    peaks = process_input(peak_file, peak_type, option, boundary)
    decomposed_peaks = decompose_peaks(peaks)
    output = find_nearest(decomposed_peaks, species, num_features, ref_dir, up_bound, down_bound)
    if output_type == 'xlsx':
        write_to_excel(output, output_name, out_dir)
    elif output_type == 'csv':
        write_to_csv(output, output_name, out_dir)
    else:
        raise ValueError('Invalid output type')

def decompose_peaks(peaks: pl.DataFrame) -> dict:
    '''
    Decompose peaks by chromosome.

    Parameters:
    peaks (pl.DataFrame): Polars DataFrame containing peak information.

    Returns:
    decomposed_peaks (dict): Dictionary containing keys with chromosome number
                             mapped to Polars DataFrames with peaks on that chromosome

    Outputs:
    None
    '''

    return {'chr' + str(name[0]) if 'chr' not in str(name[0]) else str(name[0]): 
            group for name, group in peaks.group_by(['chr'])}

def find_nearest(decomposed_peaks: dict, 
                 species: str, 
                 num_features: int, 
                 ref_dir: str, 
                 up_bound: int, 
                 down_bound: int) -> pd.DataFrame:
    '''
    Find the nearest genes for a given list of peaks. Place these in a Pandas DataFrame

    Parameters:
    decomposed_peaks (dict): Dictionary containing keys with chromosome number
                             mapped to Polars DataFrames with peaks on that chromosome
    species (str): Species of the reference genome.
    num_features (int): Number of nearest features to find.
    ref_dir (str): Directory containing decomposed reference data.
    up_bound (int): Maximum allowed distance between peak and upstream feature.
    down_bound (int): Maximum allowed distance between peak and downstream feature.

    Returns:
    output (pd.DataFrame): Pandas DataFrame containing peak data, the nearest k genes for each peak, 
    and the distance between those genes and the peak.

    Outputs:
    None
    '''

    output = pl.DataFrame()
    for key in decomposed_peaks.keys():
        starts = pl.read_csv(os.path.join(ref_dir, species, 'gene', key) + '_start.csv')
        ends = pl.read_csv(os.path.join(ref_dir, species, 'gene', key) + '_end.csv')
        output = pl.concat([output, get_nearest_features(decomposed_peaks[key], 'gene_name', starts, ends,
                                                         up_bound, down_bound, num_features)])
    
    output = output.to_pandas()
    output = output.sort_values(by=['chr', 'start'])

    return output
 
def write_to_excel(output: pd.DataFrame, 
                   output_name: str, 
                   out_dir: str) -> None:
    '''
    Write output Pandas DataFrame to an Excel sheet

    Parameters:
    output (pd.DataFrame): Pandas DataFrame containing peak data, the nearest k genes for each peak, 
                           and the distance between those genes and the peak.
    output_name (str): Name for output file.
    out_dir (str): Directory to output file.

    Returns:
    None

    Outputs:
    Excel sheet containing peak data, the nearest k genes for each peak, and the distance
    between those genes and the peak.
    '''

    if not os.path.exists(out_dir):
        os.mkdir(out_dir)

    with pd.ExcelWriter(os.path.join(out_dir, output_name) + '.xlsx', engine='openpyxl') as writer:
    
        output.to_excel(writer, sheet_name='Sheet1', index=False)

        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        for row_num in range(2, len(output) + 2):
            if row_num % 2 == 0:
                for col_num in range(1, output.shape[1] + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
            else:
                for col_num in range(1, output.shape[1] + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width


        unique_chr_values = output['chr'].unique()

        filters = worksheet.auto_filter
        filters.ref = "B1:B" + str(len(output) + 1)
        col = FilterColumn(colId=0)
        col.filters = Filters(filter=unique_chr_values.tolist())
        filters.filterColumn.append(col)

        workbook.save(os.path.join(out_dir, output_name) + '.xlsx')

def write_to_csv(output: pd.DataFrame, 
                 output_name: str,
                 out_dir: str) -> None:
    '''
    Write output Pandas DataFrame to an CSV file

    Parameters:
    output (pd.DataFrame): Pandas DataFrame containing peak data, the nearest k genes for each peak, 
                           and the distance between those genes and the peak.
    output_name (str): Name for output file.
    out_dir (str): Directory to output file.

    Returns:
    None

    Outputs:
    CSV file containing peak data, the nearest k genes for each peak, and the distance
    between those genes and the peak.
    '''

    if not os.path.exists(out_dir):
        os.mkdir(out_dir)

    output.to_csv(os.path.join(out_dir, output_name) + '.csv', index=False)
    
# data information
# data_dir = "test/"
# ref_dir = "reference/"

# # function parameters
# num_peaks_cutoff = None 
# num_nearest_features = 3

# option = "native_peak_boundaries"
# boundary = None

# file = "test_MACS2.bed"
# peak2gene(data_dir + file, 'MACS2', "test", 3, ref_dir, file[:-4], 'results')
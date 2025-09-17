# ------------------------------------------------------------------------------
#                        __   _____                  __
#      ____  ___  ____ _/ /__/ ___/_________  __  __/ /_
#     / __ \/ _ \/ __ `/ //_/\__ \/ ___/ __ \/ / / / __/
#    / /_/ /  __/ /_/ / ,<  ___/ / /__/ /_/ / /_/ / /_
#   / .___/\___/\__,_/_/|_|/____/\___/\____/\__,_/\__/
#  /_/
#
# Copyrigh 2025 GNU AFFERO GENERAL PUBLIC LICENSE
# Alexander L. Lin, Lana A. Cartailler, Jean-Philippe Cartailler
# https://github.com/vandydata/peakScout
#
# ------------------------------------------------------------------------------

import pandas as pd
import os
from openpyxl.styles import PatternFill
from openpyxl.worksheet.filters import FilterColumn, Filters
from openpyxl.utils import get_column_letter


def write_to_excel(output: pd.DataFrame, output_name: str, out_dir: str) -> None:
    """
    Write output Pandas DataFrame to an Excel sheet

    Parameters:
    output (pd.DataFrame): Pandas DataFrame containing peak data, the nearest k genes for each peak,
                           and the distance between those genes and the peak.
    output_name (str): Name for output file.
    out_dir (str): Directory to output file.

    Returns:
    None

    Outputs:
    Excel sheet containing peak data, the nearest k genes for each peak, and the distance
    between those genes and the peak.
    """

    if not os.path.exists(out_dir):
        os.mkdir(out_dir)

    with pd.ExcelWriter(
        os.path.join(out_dir, output_name) + ".xlsx", engine="openpyxl"
    ) as writer:

        output.to_excel(writer, sheet_name="Sheet1", index=False)

        workbook = writer.book
        worksheet = writer.sheets["Sheet1"]

        for row_num in range(2, len(output) + 2):
            if row_num % 2 == 0:
                for col_num in range(1, output.shape[1] + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.fill = PatternFill(
                        start_color="E6E6E6", end_color="E6E6E6", fill_type="solid"
                    )
            else:
                for col_num in range(1, output.shape[1] + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.fill = PatternFill(
                        start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
                    )

        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            worksheet.column_dimensions[get_column_letter(column[0].column)].width = (
                adjusted_width
            )

        chr_col_idx = None
        for idx, cell in enumerate(worksheet[1], start=1):
            if cell.value == "chr":
                chr_col_idx = idx
                break

        if chr_col_idx is not None:
            col_letter = get_column_letter(chr_col_idx)

            unique_chr_values = output["chr"].unique()

            filters = worksheet.auto_filter
            filters.ref = f"{col_letter}1:{col_letter}{len(output) + 1}"
            col = FilterColumn(colId=chr_col_idx - 1)
            col.filters = Filters(filter=unique_chr_values.tolist())
            filters.filterColumn.append(col)

        workbook.save(os.path.join(out_dir, output_name) + ".xlsx")


def write_to_csv(output: pd.DataFrame, output_name: str, out_dir: str) -> None:
    """
    Write output Pandas DataFrame to an CSV file

    Parameters:
    output (pd.DataFrame): Pandas DataFrame containing peak data, the nearest k genes for each peak,
                           and the distance between those genes and the peak.
    output_name (str): Name for output file.
    out_dir (str): Directory to output file.

    Returns:
    None

    Outputs:
    CSV file containing peak data, the nearest k genes for each peak, and the distance
    between those genes and the peak.
    """

    if not os.path.exists(out_dir):
        os.mkdir(out_dir)

    output.to_csv(os.path.join(out_dir, output_name) + ".csv", index=False)
